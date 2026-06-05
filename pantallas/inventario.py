import csv
import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from constantes import *
from datos import GestorArchivos
from entidades import InventarioCorte, InventarioItem
from .base import FrameBase
from .lazy import FrameMenuAdmin


COLUMNAS = ("nombre", "categoria", "total", "danadas", "disponibles", "costo_unitario", "reposicion", "notas")
ENCABEZADOS = {
    "nombre": "Articulo",
    "categoria": "Categoria",
    "total": "Total",
    "danadas": "Danadas",
    "disponibles": "Disponibles",
    "costo_unitario": "Costo unit.",
    "reposicion": "Reposicion",
    "notas": "Notas",
}

COLUMNAS_REQUERIDAS = {
    "articulo": {"articulo", "nombre", "objeto"},
    "categoria": {"categoria"},
    "total": {"total"},
    "danadas": {"danadas", "danados", "dañadas"},
    "costo_unitario": {"costo_unitario", "costo"},
    "notas": {"notas"},
}


def normalizar_clave(valor):
    return str(valor or "").strip().lower().replace(" ", "_").replace("ñ", "n")


class FrameInventario(FrameBase):
    def __init__(self, master, **kwargs):
        self.items = []
        self.indice_editando = None
        super().__init__(master, **kwargs)

    def configurar(self):
        tk.Label(self, text="Inventario y costos", font=("Arial", 18, "bold"), bg=BG, fg=TXT).pack(pady=(14, 2))
        tk.Label(self, text="Registra un corte de inventario, importa archivos y compara diferencias.", font=("Arial", 10), bg=BG, fg="#555").pack()

        frame_main = tk.Frame(self, bg=BG)
        frame_main.pack(fill="both", expand=True, padx=14, pady=10)

        self.crear_formulario(frame_main)
        self.crear_tabla(frame_main)
        self.crear_acciones(frame_main)
        self.actualizar_resumen()

    def crear_formulario(self, parent):
        frame = tk.Frame(parent, bg="white", highlightbackground="#D1D5DB", highlightthickness=1)
        frame.pack(fill="x", pady=(0, 10))

        campos = [
            ("Articulo", "nombre", 18),
            ("Categoria", "categoria", 14),
            ("Total", "total", 8),
            ("Danadas", "danadas", 8),
            ("Costo unit.", "costo", 10),
            ("Notas", "notas", 22),
        ]
        self.entries = {}
        for col, (label, key, width) in enumerate(campos):
            tk.Label(frame, text=label, font=("Arial", 9, "bold"), bg="white", fg=TXT).grid(row=0, column=col, padx=6, pady=(8, 2), sticky="w")
            entry = tk.Entry(frame, width=width)
            entry.grid(row=1, column=col, padx=6, pady=(0, 8), sticky="w")
            self.entries[key] = entry

        tk.Button(frame, text="Cargar objeto", command=self.agregar_item, bg=BTN, fg="white",
                  activebackground=BTN, activeforeground="white", relief="flat", width=16, cursor="hand2").grid(row=1, column=len(campos), padx=8, pady=(0, 8))
        tk.Button(frame, text="Actualizar seleccionado", command=self.actualizar_item, bg=BTN2, fg="white",
                  activebackground=BTN2, activeforeground="white", relief="flat", width=20, cursor="hand2").grid(row=1, column=len(campos) + 1, padx=8, pady=(0, 8))

    def crear_tabla(self, parent):
        frame = tk.Frame(parent, bg=BG)
        frame.pack(fill="both", expand=True)

        self.tabla = ttk.Treeview(frame, columns=COLUMNAS, show="headings", height=12)
        for col in COLUMNAS:
            self.tabla.heading(col, text=ENCABEZADOS[col])
            ancho = 90
            if col in ("nombre", "notas"):
                ancho = 150
            self.tabla.column(col, width=ancho, anchor="center")

        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=scroll.set)
        self.tabla.bind("<<TreeviewSelect>>", self.cargar_seleccion_en_formulario)
        self.tabla.bind("<Double-1>", self.cargar_seleccion_en_formulario)
        self.tabla.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

    def crear_acciones(self, parent):
        frame = tk.Frame(parent, bg=BG)
        frame.pack(fill="x", pady=10)

        acciones = [
            ("Guardar corte", self.guardar_corte, BTN),
            ("Cargar CSV/XLSX", self.cargar_archivo, BTN2),
            ("Comparar archivo", self.comparar_archivo, "#0F766E"),
            ("Eliminar seleccionado", self.eliminar_seleccionado, "#B45309"),
            ("Volver", lambda: self.volver(FrameMenuAdmin), "#777"),
        ]
        for texto, comando, color in acciones:
            tk.Button(frame, text=texto, command=comando, bg=color, fg="white", activebackground=color,
                      activeforeground="white", relief="flat", width=18, height=2, cursor="hand2").pack(side="left", padx=5)

        self.lbl_resumen = tk.Label(parent, text="", font=("Arial", 10, "bold"), bg=BG, fg=TXT)
        self.lbl_resumen.pack(anchor="w", pady=(2, 0))

        formato = "Formato aceptado: articulo/nombre, categoria, total, danadas, costo_unitario/costo, notas."
        tk.Label(parent, text=formato, font=("Arial", 9), bg=BG, fg="#555").pack(anchor="w")

    def agregar_item(self):
        item = self.item_desde_formulario()
        if item is None:
            return

        self.items.append(item)
        self.insertar_item_tabla(item)
        self.actualizar_resumen()
        self.limpiar_formulario()

    def item_desde_formulario(self):
        try:
            item = InventarioItem(
                self.entries["nombre"].get().strip(),
                self.entries["categoria"].get().strip(),
                self.entries["total"].get().strip() or 0,
                self.entries["danadas"].get().strip() or 0,
                self.entries["costo"].get().strip() or 0,
                self.entries["notas"].get().strip(),
            )
        except ValueError:
            messagebox.showerror("Inventario", "Total, danadas y costo unitario deben ser numeros.")
            return

        if not item.nombre:
            messagebox.showwarning("Inventario", "Ingresa el nombre del articulo.")
            return None

        return item

    def limpiar_formulario(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
        self.indice_editando = None

    def cargar_seleccion_en_formulario(self, event=None):
        seleccion = self.tabla.selection()
        if not seleccion:
            return
        indice = self.tabla.index(seleccion[0])
        if indice >= len(self.items):
            return
        self.indice_editando = indice
        item = self.items[indice]

        valores = {
            "nombre": item.nombre,
            "categoria": item.categoria,
            "total": item.total,
            "danadas": item.danadas,
            "costo": item.costo_unitario,
            "notas": item.notas,
        }
        for key, entry in self.entries.items():
            entry.delete(0, tk.END)
            entry.insert(0, str(valores.get(key, "")))

    def actualizar_item(self):
        if self.indice_editando is None:
            seleccion = self.tabla.selection()
            if not seleccion:
                messagebox.showinfo("Inventario", "Selecciona un articulo para actualizar.")
                return
            self.indice_editando = self.tabla.index(seleccion[0])

        if self.indice_editando >= len(self.items):
            messagebox.showerror("Inventario", "No se encontro el articulo seleccionado.")
            self.indice_editando = None
            return

        item = self.item_desde_formulario()
        if item is None:
            return
        self.items[self.indice_editando] = item
        self.refrescar_tabla()
        self.limpiar_formulario()

    def insertar_item_tabla(self, item):
        self.tabla.insert("", "end", values=(
            item.nombre,
            item.categoria,
            item.total,
            item.danadas,
            item.disponibles,
            f"{item.costo_unitario:.2f}",
            f"{item.costo_reposicion:.2f}",
            item.notas,
        ))

    def refrescar_tabla(self):
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        for item in self.items:
            self.insertar_item_tabla(item)
        self.actualizar_resumen()

    def actualizar_resumen(self):
        total_articulos = len(self.items)
        total_reposicion = sum(item.costo_reposicion for item in self.items)
        self.lbl_resumen.config(text=f"Articulos: {total_articulos} | Reposicion estimada: ${total_reposicion:.2f}")

    def eliminar_seleccionado(self):
        seleccion = self.tabla.selection()
        if not seleccion:
            messagebox.showinfo("Inventario", "Selecciona un articulo para eliminar.")
            return
        indices = [self.tabla.index(item_id) for item_id in seleccion]
        for indice in sorted(indices, reverse=True):
            del self.items[indice]
        self.refrescar_tabla()
        self.limpiar_formulario()

    def guardar_corte(self):
        if not self.items:
            messagebox.showwarning("Inventario", "Agrega al menos un articulo antes de guardar.")
            return
        fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        corte = InventarioCorte(fecha, list(self.items))
        ruta_csv = GestorArchivos.guardar_corte_inventario(corte)
        cantidad = len(self.items)
        self.items = []
        self.refrescar_tabla()
        self.limpiar_formulario()
        messagebox.showinfo("Inventario", f"Corte guardado con {cantidad} articulos.\nCSV generado:\n{ruta_csv}")

    def cargar_archivo(self):
        ruta = filedialog.askopenfilename(
            title="Cargar inventario",
            filetypes=[("CSV o Excel", "*.csv *.xlsx"), ("CSV", "*.csv"), ("Excel", "*.xlsx")]
        )
        if not ruta:
            return
        try:
            self.items = self.leer_items_archivo(ruta)
            self.refrescar_tabla()
            self.limpiar_formulario()
            messagebox.showinfo("Inventario", f"Se cargaron {len(self.items)} articulos.")
        except Exception as exc:
            messagebox.showerror("Inventario", f"No se pudo cargar el archivo:\n{exc}")

    def comparar_archivo(self):
        if not self.items:
            messagebox.showwarning("Comparacion", "Carga o captura primero el inventario actual.")
            return
        ruta = filedialog.askopenfilename(
            title="Comparar contra inventario",
            filetypes=[("CSV o Excel", "*.csv *.xlsx"), ("CSV", "*.csv"), ("Excel", "*.xlsx")]
        )
        if not ruta:
            return
        try:
            otros_items = self.leer_items_archivo(ruta)
        except Exception as exc:
            messagebox.showerror("Comparacion", f"No se pudo cargar el archivo:\n{exc}")
            return
        self.mostrar_comparacion(otros_items)

    def leer_items_archivo(self, ruta):
        path = Path(ruta)
        if path.suffix.lower() == ".csv":
            return self.leer_csv(path)
        if path.suffix.lower() == ".xlsx":
            return self.leer_xlsx(path)
        raise ValueError("Solo se aceptan archivos .csv o .xlsx")

    def leer_csv(self, path):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return self.items_desde_filas(reader)

    def leer_xlsx(self, path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Para cargar Excel se necesita openpyxl. Usa CSV mientras tanto.") from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalizar_clave(c) for c in rows[0]]
        dict_rows = []
        for row in rows[1:]:
            dict_rows.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
        return self.items_desde_filas(dict_rows)

    def items_desde_filas(self, filas):
        items = []
        filas = [dict(fila) for fila in filas]
        if not filas:
            return []
        columnas = {normalizar_clave(k) for fila in filas for k in fila.keys()}
        faltantes = []
        for nombre_grupo, opciones in COLUMNAS_REQUERIDAS.items():
            if not (columnas & opciones):
                faltantes.append(nombre_grupo)
        if faltantes:
            raise ValueError("Faltan columnas requeridas: " + ", ".join(faltantes))

        for fila in filas:
            datos = {normalizar_clave(k): v for k, v in fila.items()}
            nombre = datos.get("articulo") or datos.get("nombre") or datos.get("objeto")
            if not nombre:
                raise ValueError("Hay una fila sin articulo/nombre.")
            items.append(InventarioItem(
                nombre,
                datos.get("categoria", ""),
                datos.get("total", 0),
                datos.get("danadas") or datos.get("danados") or datos.get("dañadas") or 0,
                datos.get("costo_unitario") or datos.get("costo") or 0,
                datos.get("notas", ""),
            ))
        return items

    def mostrar_comparacion(self, otros_items):
        actual = {normalizar_clave(item.nombre): item for item in self.items}
        anterior = {normalizar_clave(item.nombre): item for item in otros_items}
        nombres = sorted(set(actual) | set(anterior))
        resultados = []
        costo_total_reposicion = 0.0

        ventana = tk.Toplevel(self)
        ventana.title("Comparacion de inventario")
        ventana.geometry("720x420")
        ventana.configure(bg=BG)

        texto = tk.Text(ventana, wrap="word", bg="white", fg=TXT)
        texto.pack(fill="both", expand=True, padx=12, pady=12)
        texto.insert("end", "Comparacion preparada\n\n")
        for nombre in nombres:
            item_actual = actual.get(nombre)
            item_anterior = anterior.get(nombre)
            total_actual = item_actual.total if item_actual else 0
            total_anterior = item_anterior.total if item_anterior else 0
            diff = total_actual - total_anterior
            costo_unitario = 0.0
            if item_actual and item_actual.costo_unitario:
                costo_unitario = item_actual.costo_unitario
            elif item_anterior:
                costo_unitario = item_anterior.costo_unitario

            faltantes = abs(diff) if diff < 0 else 0
            costo_reposicion = faltantes * costo_unitario
            costo_total_reposicion += costo_reposicion

            if diff < 0:
                estado = f"faltan {abs(diff)}"
            elif diff > 0:
                estado = f"sobran {diff}"
            else:
                estado = "sin cambio"
            etiqueta = item_actual.nombre if item_actual else item_anterior.nombre
            linea_costo = f" | reposicion: ${costo_reposicion:.2f}" if costo_reposicion > 0 else ""
            texto.insert("end", f"{etiqueta}: actual {total_actual}, comparado {total_anterior}, {estado}{linea_costo}\n")
            resultados.append({
                "articulo": etiqueta,
                "total_actual": total_actual,
                "total_comparado": total_anterior,
                "diferencia": diff,
                "faltantes": faltantes,
                "costo_unitario": costo_unitario,
                "costo_reposicion": costo_reposicion,
            })

        data = {
            "fecha": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "costo_reposicion_total": costo_total_reposicion,
            "articulos": resultados,
        }
        GestorArchivos.guardar_comparacion_inventario(data)
        resumen_reposicion = GestorArchivos.resumen_reposicion_inventario()

        texto.insert("end", "\n")
        texto.insert("end", f"TOTAL GENERAL A REPONER EN ESTA COMPARACION: ${costo_total_reposicion:.2f}\n")
        if resumen_reposicion["total_comparaciones"] >= 3:
            texto.insert("end", f"Promedio historico del salon ({resumen_reposicion['total_comparaciones']} comparaciones): ${resumen_reposicion['promedio']:.2f}\n")
            texto.insert("end", "Este promedio sirve como referencia para apartar dinero por periodo.\n")
        else:
            faltantes_promedio = 3 - resumen_reposicion["total_comparaciones"]
            texto.insert("end", f"Promedio historico pendiente: faltan {faltantes_promedio} comparacion(es) para calcularlo con base minima.\n")
            texto.insert("end", "El total de esta comparacion ya quedo guardado para ese promedio futuro.\n")
        texto.config(state="disabled")
